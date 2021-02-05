import openpyxl
import mysql.connector

wb = openpyxl.load_workbook('C:/Users/SkySystem/Desktop/Info.xlsx')
sheet = wb['Sheet1']

sql = mysql.connector.connect(user='AM80', password='@Ali2001', host='127.0.0.1',
                              database='Python')
operator = sql.cursor()

cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q',
        'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

mycols = cols[:sheet.max_column]
requests = list()
req_str = str()

for i in range(2, sheet.max_row+1):
    requests = []
    req_str = ''
    for m in mycols:
        requests.append(str(sheet[m+str(i)].value))
    print(requests)
    for m in requests:
        req_str = ', '.join([req_str, "'" + m + "'"])
    print(req_str[2:])
    operator.execute("INSERT INTO Learning VALUES (" + req_str[2:] + ");")

sql.commit()